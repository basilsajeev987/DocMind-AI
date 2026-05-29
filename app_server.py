import os
import re
import sys
import json
import time
import math
import base64
import tempfile
import traceback
from dataclasses import dataclass
from typing import List, Dict, Tuple, Optional

import numpy as np  # kept (not required for keyword RAG)

# Ensure UTF-8 logs on Windows terminals
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

from src.llm.llm_pipeline import LLMPipeline

# =========================
# CONFIG
# =========================
MAX_FILES = 200
MAX_CHUNKS = 5000
MAX_FILE_BYTES = 100 * 1024 * 1024   # ✅ 100MB per file

CHUNK_CHARS = 900
CHUNK_OVERLAP = 150
TOP_K = 5

# If your PDFs are scanned images, set OCR_ENABLED=True and install OCR deps.
OCR_ENABLED = True
OCR_DPI = 250
OCR_MAX_PAGES = 25  # safety for huge PDFs

ALLOWED_EXT = {
    ".txt", ".md", ".rst",
    ".py", ".js", ".ts", ".tsx", ".jsx",
    ".json", ".yml", ".yaml",
    ".html", ".css",
    ".java", ".c", ".cpp", ".h",
    ".go", ".rs", ".php", ".sql",
    ".log",
    ".pdf",
    ".docx",
    ".xlsx",
}

_token_re = re.compile(r"[A-Za-z0-9_]+")


# =========================
# Logging (IMPORTANT: stderr only)
# =========================
def log_err(*args):
    try:
        sys.stderr.write("[RAG] " + " ".join(str(a) for a in args) + "\n")
        sys.stderr.flush()
    except Exception:
        pass


# =========================
# Helpers
# =========================
def _safe_str(x) -> str:
    if x is None:
        return ""
    if isinstance(x, str):
        return x
    try:
        return json.dumps(x, ensure_ascii=False)
    except Exception:
        return str(x)

def tokenize(text: str) -> List[str]:
    return [t.lower() for t in _token_re.findall(text or "")]

def list_files_in_folder(folder: str, exts=(".pdf", ".docx", ".xlsx", ".txt", ".md")):
    files = []
    for root, _, names in os.walk(folder):
        for n in names:
            if n.lower().endswith(exts):
                files.append(os.path.join(root, n))
    return files


# =========================
# File text extraction
# =========================
def extract_pdf_text_pymupdf(path: str) -> str:
    import fitz  # PyMuPDF
    doc = fitz.open(path)
    try:
        out = []
        for page in doc:
            t = page.get_text("text") or ""
            if t.strip():
                out.append(t)
        return "\n".join(out).strip()
    finally:
        try:
            doc.close()
        except Exception:
            pass

def extract_pdf_text_pypdf(path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    out = []
    for page in reader.pages:
        t = page.extract_text() or ""
        if t.strip():
            out.append(t)
    return "\n".join(out).strip()

def extract_pdf_text_ocr(path: str) -> str:
    # OCR fallback (scanned PDFs)
    # Requires: pip install pytesseract pdf2image pillow
    # Requires Tesseract installed on Windows
    from pdf2image import convert_from_path
    import pytesseract

    images = convert_from_path(path, dpi=OCR_DPI, first_page=1, last_page=OCR_MAX_PAGES)
    out = []
    for img in images:
        out.append(pytesseract.image_to_string(img) or "")
    return "\n".join(out).strip()

def extract_docx_text(path: str) -> str:
    from docx import Document
    doc = Document(path)
    parts = []
    for p in doc.paragraphs:
        if p.text and p.text.strip():
            parts.append(p.text.strip())
    return "\n".join(parts).strip()

def extract_xlsx_text(path: str) -> str:
    # XLSX can be huge; keep it safe
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    out = []
    try:
        for sheet in wb.worksheets:
            out.append(f"--- Sheet: {sheet.title} ---")
            row_count = 0
            for row in sheet.iter_rows(values_only=True):
                row_count += 1
                # safety: limit rows per sheet
                if row_count > 5000:
                    out.append("[...truncated rows...]")
                    break
                row_text = " | ".join(str(c) for c in row if c is not None)
                if row_text.strip():
                    out.append(row_text)
        return "\n".join(out).strip()
    finally:
        try:
            wb.close()
        except Exception:
            pass

def read_file_text(path: str) -> str:
    """
    Extract text from:
      - PDF  : PyMuPDF -> pypdf -> OCR(optional)
      - DOCX : python-docx
      - XLSX : openpyxl
      - others: plain text
    """
    ext = os.path.splitext(path)[1].lower()

    # ---------- PDF ----------
    if ext == ".pdf":
        # 1) PyMuPDF
        try:
            t = extract_pdf_text_pymupdf(path)
            if t.strip():
                return t
        except Exception as e:
            log_err("PyMuPDF failed:", os.path.basename(path), "|", repr(e))

        # 2) pypdf fallback
        try:
            t = extract_pdf_text_pypdf(path)
            if t.strip():
                return t
        except Exception as e:
            log_err("pypdf failed:", os.path.basename(path), "|", repr(e))

        # 3) OCR optional
        if OCR_ENABLED:
            try:
                t = extract_pdf_text_ocr(path)
                if t.strip():
                    return t
            except Exception as e:
                log_err("OCR failed:", os.path.basename(path), "|", repr(e))

        # If all failed
        return ""

    # ---------- DOCX ----------
    if ext == ".docx":
        try:
            return extract_docx_text(path)
        except Exception as e:
            log_err("DOCX failed:", os.path.basename(path), "|", repr(e))
            return ""

    # ---------- XLSX ----------
    if ext == ".xlsx":
        try:
            return extract_xlsx_text(path)
        except Exception as e:
            log_err("XLSX failed:", os.path.basename(path), "|", repr(e))
            return ""

    # ---------- TEXT / CODE ----------
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read().strip()
    except Exception as e:
        log_err("Text read failed:", os.path.basename(path), "|", repr(e))
        return ""


# =========================
# Keyword RAG
# =========================
@dataclass
class Chunk:
    path: str
    idx: int
    text: str
    tokens: List[str]


class RagIndex:
    def __init__(self):
        self.folder: str = ""
        self.chunks: List[Chunk] = []
        self.file_count: int = 0
        self.built_at: float = 0.0
        self.skipped: Dict[str, int] = {}

    def clear(self):
        self.folder = ""
        self.chunks = []
        self.file_count = 0
        self.built_at = 0.0
        self.skipped = {}

    def _skip(self, reason: str):
        self.skipped[reason] = self.skipped.get(reason, 0) + 1

    def build(self, folder: str):
        folder = os.path.abspath(folder)
        self.clear()
        self.folder = folder

        chunks: List[Chunk] = []
        file_count = 0
        total_chunks = 0

        for root, _, files in os.walk(folder):
            for fn in files:
                ext = os.path.splitext(fn)[1].lower()
                if ext not in ALLOWED_EXT:
                    self._skip("ext_not_allowed")
                    continue

                path = os.path.join(root, fn)

                # file size guard
                try:
                    if os.path.getsize(path) > MAX_FILE_BYTES:
                        self._skip("too_large")
                        continue
                except OSError:
                    self._skip("size_error")
                    continue

                # extract text
                text = read_file_text(path)
                if not text.strip():
                    self._skip("no_text_extracted")
                    continue

                file_count += 1
                if file_count > MAX_FILES:
                    self._skip("max_files_reached")
                    break

                start = 0
                idx = 0
                while start < len(text):
                    end = min(len(text), start + CHUNK_CHARS)
                    chunk_text = text[start:end]
                    toks = tokenize(chunk_text)

                    if toks:
                        chunks.append(Chunk(path=path, idx=idx, text=chunk_text, tokens=toks))
                        total_chunks += 1
                        if total_chunks >= MAX_CHUNKS:
                            self._skip("max_chunks_reached")
                            break

                    idx += 1
                    start = end - CHUNK_OVERLAP
                    if start < 0:
                        start = 0
                    if end == len(text):
                        break

                if total_chunks >= MAX_CHUNKS:
                    break

            if file_count > MAX_FILES or total_chunks >= MAX_CHUNKS:
                break

        self.chunks = chunks
        self.file_count = file_count
        self.built_at = time.time()

        log_err("Build done. files=", self.file_count, "chunks=", len(self.chunks), "skipped=", self.skipped)

    def retrieve(self, query: str, k: int = TOP_K) -> List[Tuple[float, Chunk]]:
        qtokens = tokenize(query)
        if not qtokens or not self.chunks:
            return []

        df: Dict[str, int] = {t: 0 for t in set(qtokens)}
        for ch in self.chunks:
            st = set(ch.tokens)
            for t in df.keys():
                if t in st:
                    df[t] += 1

        N = max(1, len(self.chunks))
        idf = {t: math.log((N + 1) / (df[t] + 1)) + 1.0 for t in df.keys()}

        scored: List[Tuple[float, Chunk]] = []
        qset = set(qtokens)

        for ch in self.chunks:
            if not ch.tokens:
                continue

            tok_counts: Dict[str, int] = {}
            for t in ch.tokens:
                if t in qset:
                    tok_counts[t] = tok_counts.get(t, 0) + 1

            if not tok_counts:
                continue

            tfidf = 0.0
            for t, c in tok_counts.items():
                tfidf += (1.0 + math.log(c)) * idf.get(t, 1.0)

            score = tfidf / (1.0 + math.log(len(ch.tokens) + 1))
            scored.append((score, ch))

        scored.sort(key=lambda x: x[0], reverse=True)
        return scored[:k]


def format_snips(results: List[Tuple[float, Chunk]], base_folder: str) -> List[str]:
    snips = []
    for score, ch in results:
        rel = os.path.relpath(ch.path, base_folder)
        snips.append(f"[{rel} :: chunk {ch.idx}]\n{ch.text.strip()}\n")
    return snips

def format_sources(results: List[Tuple[float, Chunk]], base_folder: str) -> List[dict]:
    sources = []
    for score, ch in results:
        rel = os.path.relpath(ch.path, base_folder)
        sources.append({
            "path": ch.path,
            "rel": rel,
            "chunk": ch.idx,
            "score": float(score),
        })
    return sources


def build_prompt(system_prompt: str, history: list, user_text: str, rag_snips: List[str]) -> str:
    system_prompt = (system_prompt or "").strip()
    trimmed = history[-10:] if isinstance(history, list) else []

    parts: List[str] = []
    if system_prompt:
        parts.append("### System")
        parts.append(system_prompt)
        parts.append("")

    if rag_snips:
        parts.append("### Context (from Basil's local files)")
        parts.append("Use this context if relevant. If not found in context, say you are not sure.")
        parts.append("")
        parts.extend(rag_snips)
        parts.append("")

    parts.append("### Conversation")
    for m in trimmed:
        if not isinstance(m, dict):
            continue
        role = (m.get("role") or "").lower()
        text = (m.get("text") or "").strip()
        if not text:
            continue
        if role == "user":
            parts.append(f"User: {text}")
        elif role == "assistant":
            parts.append(f"Assistant: {text}")

    parts.append(f"User: {user_text.strip()}")
    parts.append("Assistant:")
    return "\n".join(parts)


# =========================
# Offline ASR (faster-whisper)
# =========================
_asr_model = None

def asr_transcribe_wav_b64(wav_b64: str, model_size: str = "tiny") -> str:
    global _asr_model
    from faster_whisper import WhisperModel

    if _asr_model is None:
        _asr_model = WhisperModel(model_size, device="cpu", compute_type="int8")

    audio_bytes = base64.b64decode(wav_b64)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as f:
        tmp_path = f.name
        f.write(audio_bytes)

    try:
        segments, _info = _asr_model.transcribe(tmp_path, vad_filter=True)
        parts = []
        for seg in segments:
            if seg.text:
                parts.append(seg.text.strip())
        return " ".join(parts).strip()
    finally:
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def call_pipeline(pipe, prompt: str) -> str:
    return pipe.run(prompt)


# =========================
# MAIN LOOP
# =========================
rag = RagIndex()

def main():
    model_dir = r"C:\Users\Basil Sajeev\Desktop\per\LADYBIRD\LLM2\models\LLM\Qwen2.5-Coder-1.5B-Instruct"

    pipe = LLMPipeline(
        model_dir=model_dir,
        device="cuda",
        quantization="auto",
        max_new_tokens=120,
    )

    sys.stdout.write(json.dumps({"type": "ready"}) + "\n")
    sys.stdout.flush()

    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue

        try:
            msg = json.loads(line)
        except Exception:
            sys.stdout.write(json.dumps({"type": "error", "message": "Bad JSON input"}) + "\n")
            sys.stdout.flush()
            continue

        mtype = msg.get("type")

        if mtype == "exit":
            break

        # ---------- ASR ----------
        if mtype == "asr":
            try:
                wav_b64 = (msg.get("wav_b64") or "").strip()
                model_size = (msg.get("model_size") or "tiny").strip()
                request_id = msg.get("request_id") or f"asr_{int(time.time()*1000)}"
                if not wav_b64:
                    sys.stdout.write(json.dumps({"type": "asr", "request_id": request_id, "text": ""}) + "\n")
                    sys.stdout.flush()
                    continue

                text = asr_transcribe_wav_b64(wav_b64, model_size=model_size)
                sys.stdout.write(json.dumps({"type": "asr", "request_id": request_id, "text": text}) + "\n")
            except Exception as e:
                tb = traceback.format_exc()
                sys.stdout.write(json.dumps({"type": "error", "message": str(e), "traceback": tb}) + "\n")
            sys.stdout.flush()
            continue

        # ---------- PROMPT ----------
        if mtype != "prompt":
            sys.stdout.write(json.dumps({"type": "error", "message": "Unknown message type"}) + "\n")
            sys.stdout.flush()
            continue

        user_text = _safe_str(msg.get("text")).strip()
        system_prompt = _safe_str(msg.get("system")).strip()

        raw_history = msg.get("history")
        history = raw_history if isinstance(raw_history, list) else []

        clean_history = []
        for m in history:
            if not isinstance(m, dict):
                continue
            role = _safe_str(m.get("role")).lower().strip()
            text = _safe_str(m.get("text")).strip()
            if role and text:
                clean_history.append({"role": role, "text": text})
        history = clean_history

        rag_cfg = msg.get("rag") or {}
        rag_enabled = bool(rag_cfg.get("enabled"))
        rag_folder = (rag_cfg.get("folderPath") or "").strip()

        lower_q = user_text.lower()

        # ---------- List files intent (no LLM) ----------
        if any(k in lower_q for k in [
            "list pdf", "show pdf", "list file", "show file",
            "documents in folder", "files in folder", "list the files"
        ]):
            if not rag_folder:
                sys.stdout.write(json.dumps({
                    "type": "response",
                    "text": "❌ No folder selected. Please pick a folder first.",
                    "sources": []
                }, ensure_ascii=False) + "\n")
                sys.stdout.flush()
                continue

            files = list_files_in_folder(rag_folder, exts=tuple(ALLOWED_EXT))
            if not files:
                text = "📁 No supported files found in the selected folder."
            else:
                lines = ["📁 Files found:\n"]
                for f in sorted(files):
                    lines.append(f"- {os.path.basename(f)}")
                text = "\n".join(lines)

            sys.stdout.write(json.dumps({
                "type": "response",
                "text": text,
                "sources": []
            }, ensure_ascii=False) + "\n")
            sys.stdout.flush()
            continue

        if not user_text:
            sys.stdout.write(json.dumps({"type": "response", "text": ""}) + "\n")
            sys.stdout.flush()
            continue

        try:
            rag_snips: List[str] = []
            sources: List[dict] = []
            rag_meta = None

            if rag_enabled and rag_folder:
                folder_abs = os.path.abspath(rag_folder)

                if rag.folder != folder_abs or not rag.chunks:
                    t0 = time.time()
                    rag.build(folder_abs)
                    rag_meta = {
                        "folder": folder_abs,
                        "files": rag.file_count,
                        "chunks": len(rag.chunks),
                        "build_seconds": round(time.time() - t0, 2),
                        "skipped": rag.skipped,
                    }

                results = rag.retrieve(user_text, k=TOP_K)
                rag_snips = format_snips(results, folder_abs)
                sources = format_sources(results, folder_abs)

            prompt = build_prompt(system_prompt, history, user_text, rag_snips)
            if not isinstance(prompt, str):
                prompt = str(prompt)

            response = call_pipeline(pipe, prompt)

            out = {
                "type": "response",
                "text": str(response),
                "rag_used": bool(rag_enabled and rag_folder),
                "rag_mode": "keyword",
                "sources": sources,
            }
            if rag_meta:
                out["rag_meta"] = rag_meta

            sys.stdout.write(json.dumps(out, ensure_ascii=False) + "\n")
            sys.stdout.flush()

        except Exception as e:
            tb = traceback.format_exc()
            sys.stdout.write(json.dumps({
                "type": "error",
                "message": str(e),
                "traceback": tb
            }, ensure_ascii=False) + "\n")
            sys.stdout.flush()


if __name__ == "__main__":
    main()
